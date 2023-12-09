#needs double checking
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2023/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##


movie_rows = soup.findAll('tr')

#for row in movie_rows[1:6]:
    #td = row.findAll('td')
    #print(td)
    #print(td[0].text)
    #print(td[1].text)
    #print(td[5].text)
    #print(td[6].text)

wb=xl.Workbook()

ws=wb.active

ws.title = 'Box Office Report'




ws['A1'] = 'Rank'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Gross'
ws['D1'] = 'Theatres'
ws['E1'] = 'Avg. Gross / Theatre'

for x in range(1,6):
    td = movie_rows[x].findAll('td')
    #print(td)
    rank = td[0].text
    title = td[1].text
    gross = int(td[5].text.replace('$', '').replace(',',''))
    theatres = int(td[6].text.replace(',',''))

    avg = gross/theatres

    # this evaulates to A2, A3, A4...
    ws['A' + str(x + 1)] = rank
    ws['B' + str(x + 1)] = title
    ws['C' + str(x + 1)] = gross
    ws['D' + str(x + 1)] = theatres
    ws['E' + str(x + 1)] = avg


ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 25\

header_font = Font(size = 16, bold = True)

for cell in ws[1:1]:
    cell.font = header_font

for cell in ws["D:D"]:
    cell.number_format = '#,##0'

for cell in ws["C:C"]:
    cell.number_format = u'"$ "#,##0.00'

for cell in ws["E:E"]:
    cell.number_format = u'"$ "#,##0.00'


wb.save('BoxOfficeReport.xlsx')
